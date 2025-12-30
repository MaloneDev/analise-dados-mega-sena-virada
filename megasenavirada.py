from __future__ import annotations

import io
import random
import logging
import requests
from dataclasses import dataclass
from collections import Counter
from typing import List, Dict

from openpyxl import load_workbook

# =====================================================
# CONFIGURAÇÕES
# =====================================================

DOWNLOAD_URL = (
    "https://servicebus2.caixa.gov.br/"
    "portaldeloterias/api/resultados/download"
)

PARAMS = {
    "modalidade": "Mega-Sena"
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept": "*/*"
}

TOTAL_JOGOS_GERADOS = 10
DEZENAS_POR_JOGO = 6
RANGE_DEZENAS = range(1, 61)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

# =====================================================
# MODELO DE DOMÍNIO
# =====================================================

@dataclass(frozen=True)
class Concurso:
    numero: int
    dezenas: List[int]

# =====================================================
# LOADER XLSX (AUTÔNOMO)
# =====================================================

class MegaSenaXlsxLoader:
    """
    Baixa e processa automaticamente o arquivo
    mega-sena.xlsx diretamente da Caixa.
    """

    def load(self) -> List[Concurso]:
        logging.info("Baixando arquivo XLSX da Mega-Sena")

        response = requests.get(
            DOWNLOAD_URL,
            params=PARAMS,
            headers=HEADERS,
            timeout=30
        )
        response.raise_for_status()

        workbook = load_workbook(
            filename=io.BytesIO(response.content),
            data_only=True
        )

        sheet = workbook.active

        concursos: List[Concurso] = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            try:
                numero = int(row[0])
                dezenas = list(map(int, row[2:8]))

                if len(dezenas) != 6:
                    continue

                concursos.append(
                    Concurso(
                        numero=numero,
                        dezenas=dezenas
                    )
                )
            except (TypeError, ValueError):
                continue

        if not concursos:
            raise RuntimeError("Nenhum concurso válido encontrado no XLSX")

        logging.info("Total de concursos carregados: %s", len(concursos))
        return concursos

# =====================================================
# ANÁLISE ESTATÍSTICA
# =====================================================

class MegaSenaAnalyzer:
    def __init__(self, concursos: List[Concurso]):
        self.concursos = concursos
        self.todos = [d for c in concursos for d in c.dezenas]

    def frequency_map(self) -> Dict[int, int]:
        return Counter(self.todos)

    def soma_media(self) -> float:
        return round(
            sum(sum(c.dezenas) for c in self.concursos) / len(self.concursos), 2
        )

    def par_impar_ratio(self) -> Dict[str, float]:
        pares = sum(1 for n in self.todos if n % 2 == 0)
        total = len(self.todos)
        return {
            "pares": round(pares / total, 3),
            "impares": round(1 - pares / total, 3)
        }

    def top_dezenas(self, limit: int = 10) -> List[int]:
        return [n for n, _ in self.frequency_map().most_common(limit)]

# =====================================================
# GERADOR DE JOGOS
# =====================================================

class GameGenerator:
    def __init__(self, analyzer: MegaSenaAnalyzer):
        self.freq = analyzer.frequency_map()

    def gerar_jogo(self) -> List[int]:
        dezenas = list(RANGE_DEZENAS)
        pesos = [self.freq.get(d, 1) for d in dezenas]

        jogo = set()
        while len(jogo) < DEZENAS_POR_JOGO:
            jogo.add(random.choices(dezenas, weights=pesos, k=1)[0])

        jogo = sorted(jogo)

        pares = sum(1 for n in jogo if n % 2 == 0)
        if pares < 2 or pares > 4:
            return self.gerar_jogo()

        return jogo

    def gerar_jogos(self, total: int) -> List[List[int]]:
        jogos = set()
        while len(jogos) < total:
            jogos.add(tuple(self.gerar_jogo()))
        return [list(j) for j in jogos]

# =====================================================
# MAIN
# =====================================================

def main() -> None:
    try:
        loader = MegaSenaXlsxLoader()
        concursos = loader.load()

        analyzer = MegaSenaAnalyzer(concursos)

        logging.info("Soma média histórica: %s", analyzer.soma_media())
        logging.info("Par x Ímpar: %s", analyzer.par_impar_ratio())
        logging.info("Top dezenas: %s", analyzer.top_dezenas())

        generator = GameGenerator(analyzer)
        jogos = generator.gerar_jogos(TOTAL_JOGOS_GERADOS)

        print("\n🎯 JOGOS GERADOS (BASE XLSX OFICIAL)\n")
        for i, jogo in enumerate(jogos, 1):
            print(f"Jogo {i}: {jogo}")

    except Exception as exc:
        logging.error("Erro fatal: %s", exc)

if __name__ == "__main__":
    main()
